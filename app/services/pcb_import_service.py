"""
Service d'import de fichiers Excel de bilan PCB UEMOA.

Parse un fichier .xlsx contenant un bilan réglementaire et le convertit au format
`{postes: [...], totaux: {...}}` compatible avec le reste du pipeline (analyse IA,
stockage en base, affichage).

Hypothèses sur le format Excel attendu :
- Une ligne de données contient un **code PCB** (A100, A1500, P200, P1000, etc.)
  soit seul dans une cellule, soit combiné avec le libellé ("P200 - DETTES INTERBANCAIRES").
- Les valeurs numériques suivent dans les colonnes suivantes, dans l'ordre :
  N-1, Référence intermédiaire, Clôture (la colonne "Taux" est ignorée).
- Les valeurs sont en **millions de FCFA** par défaut (configurable via le paramètre `unit`).

Le parser est tolérant : il ignore les lignes vides, les en-têtes, les lignes
de totaux intermédiaires, etc. Il se concentre sur les lignes contenant un code PCB.
"""
import re
from io import BytesIO
from typing import Dict, List, Optional
from uuid import uuid4

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None  # type: ignore


# Pattern : code PCB = lettre A ou P suivie de 3 ou 4 chiffres
PCB_CODE_RE = re.compile(r'^([APap]\d{3,4})\s*[-—–:]?\s*(.*)$')


def _parse_numeric(cell_value) -> Optional[float]:
    """Convertit une valeur de cellule en float si possible, None sinon."""
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float)):
        f = float(cell_value)
        return f if f == f else None  # exclut NaN
    if isinstance(cell_value, str):
        s = cell_value.strip()
        if not s:
            return None
        # Nettoyer : espaces insécables, séparateurs de milliers, virgule décimale
        s = s.replace('\xa0', '').replace(' ', '').replace(',', '.')
        # Gérer les pourcentages (ignorés côté valeur comptable)
        if s.endswith('%'):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def parse_excel_bilan(
    file_bytes: bytes,
    unit: str = 'millions',
) -> Dict:
    """
    Parse un fichier Excel de bilan PCB UEMOA.

    Args:
        file_bytes: contenu binaire du fichier .xlsx
        unit: 'millions' (défaut) ou 'bruts' — unité des valeurs dans le fichier

    Returns:
        dict au format {
            "postes": [...],
            "postes_hierarchiques": [],
            "totaux": {"total_actif", "total_passif", "equilibre"}
        }

    Raises:
        ValueError: si le fichier ne contient aucun code PCB reconnaissable
        ImportError: si openpyxl n'est pas installé
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl n'est pas installé. Exécutez `pip install openpyxl`.")

    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    postes: List[Dict] = []
    # Collecte de diagnostic : les valeurs texte de la 1re colonne de chaque feuille
    diag_samples: List[str] = []

    # Essaie TOUTES les feuilles — certains bilans ont plusieurs onglets
    # (Actif / Passif / Hors-bilan, ou feuille de garde + données)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not row or all(c is None for c in row):
                continue

            # Recherche du code PCB dans la ligne
            code = None
            code_cell_idx = -1
            embedded_libelle: Optional[str] = None

            for idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                m = PCB_CODE_RE.match(cell_str)
                if m:
                    code = m.group(1).upper()
                    rest = m.group(2).strip() if m.group(2) else ''
                    embedded_libelle = rest if rest else None
                    code_cell_idx = idx
                    break

            if not code:
                # Collecte un échantillon pour diagnostic
                if len(diag_samples) < 15:
                    first_text = next(
                        (str(c).strip() for c in row if c is not None and str(c).strip()),
                        None,
                    )
                    if first_text and first_text not in diag_samples:
                        diag_samples.append(first_text[:60])
                continue

            # Déterminer le libellé : soit dans la cellule du code, soit dans la suivante
            libelle = embedded_libelle
            if not libelle and code_cell_idx + 1 < len(row):
                nxt = row[code_cell_idx + 1]
                if nxt is not None:
                    nxt_str = str(nxt).strip()
                    # On évite de prendre un nombre comme libellé
                    if nxt_str and _parse_numeric(nxt_str) is None:
                        libelle = nxt_str

            libelle = libelle or code  # fallback : code si pas de libellé

            # Collecter les valeurs numériques à partir de la cellule suivant le libellé
            skip = 1 if embedded_libelle else 2
            numeric_values: List[Optional[float]] = []
            for cell in row[code_cell_idx + skip:]:
                numeric_values.append(_parse_numeric(cell))

            # On attend [N-1, Référence, Clôture, (Taux ignoré)]
            non_null = [v for v in numeric_values if v is not None]
            n_1 = non_null[0] if len(non_null) >= 1 else None
            ref = non_null[1] if len(non_null) >= 2 else None
            cloture = non_null[2] if len(non_null) >= 3 else None

            # Cas 1 valeur seule : c'est la clôture
            if n_1 is not None and ref is None and cloture is None:
                cloture = n_1
                n_1 = None

            # Cas 2 valeurs : interprétées comme [N-1, Clôture] s'il n'y a pas de 3e
            if n_1 is not None and ref is not None and cloture is None:
                cloture = ref
                ref = None

            # Conversion vers XOF bruts pour le champ clôture
            def _to_bruts(v):
                if v is None:
                    return None
                return v * 1_000_000.0 if unit == 'millions' else v

            cloture_bruts = _to_bruts(cloture)
            type_poste = 'bilan_actif' if code.startswith('A') else 'bilan_passif'

            postes.append({
                'id': str(uuid4()),
                'code': code,
                'libelle': libelle,
                'type': type_poste,
                'niveau': 0,
                'parent_id': None,
                'n_1': n_1,  # millions
                'realisation_reference': ref,  # millions
                'realisation_cloture': cloture_bruts,  # XOF bruts
                'solde_brut': cloture_bruts,
                'solde': cloture_bruts,
                'gl_codes': [],
                'source': 'excel_import',
            })

    if not postes:
        sheet_names = ", ".join(ws.title for ws in wb.worksheets)
        samples_txt = " | ".join(diag_samples[:10]) if diag_samples else "(aucun texte trouvé)"
        raise ValueError(
            f"Aucun code PCB reconnaissable dans le fichier. "
            f"Feuilles examinées : {sheet_names}. "
            f"Exemples de cellules lues : {samples_txt}. "
            f"Le parser cherche des codes au format 'A100', 'P200', 'A1500', 'P1000' etc. "
            f"(lettre A ou P suivie de 3 ou 4 chiffres). "
            f"Si tes codes ont un autre format, dis-le-moi pour adapter le parser."
        )

    # Calcul des totaux à partir des postes A1500 et P1000 (si présents)
    totaux: Dict = {}
    for p in postes:
        if p['code'] == 'A1500':
            totaux['total_actif'] = p['solde_brut'] or 0
        elif p['code'] == 'P1000':
            totaux['total_passif'] = p['solde_brut'] or 0
    if 'total_actif' in totaux and 'total_passif' in totaux:
        totaux['equilibre'] = abs(totaux['total_actif'] - totaux['total_passif']) < 1_000_000

    return {
        'postes': postes,
        'postes_hierarchiques': [],
        'totaux': totaux,
    }
