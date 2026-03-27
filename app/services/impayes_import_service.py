"""
Service d'import des impayés - Factorisation de la logique de parsing
"""

import io
import csv
from typing import List, Tuple, Dict, Any, Optional
from fastapi import UploadFile

from app.schemas.impayes import LigneImpayeImport, SegmentEnum, ProduitEnum, StatutInterneEnum


# Fonctions utilitaires locales pour éviter les dépendances croisées
def safe_cell_value(cell):
    """Convertit de manière sécurisée la valeur d'une cellule en string"""
    try:
        # Essayer d'abord de lire la valeur formatée (display_value) pour préserver le texte exact
        if hasattr(cell, 'display_value') and cell.display_value is not None:
            return str(cell.display_value)
        # Sinon, lire la valeur brute
        if cell.value is None:
            return None
        return str(cell.value)
    except Exception:
        return None


def normalize_column_name(name):
    """Normalise un nom de colonne pour la comparaison (insensible à la casse, sans espaces, caractères spéciaux)"""
    if not name:
        return ""
    # Mettre en minuscules, enlever les accents, remplacer les caractères spéciaux par des underscores
    import unicodedata
    normalized = unicodedata.normalize('NFKD', str(name))
    normalized = ''.join(c for c in normalized if not unicodedata.combining(c))
    # Remplacer tout ce qui n'est pas alphanumérique par underscore
    normalized = ''.join(c if c.isalnum() else '_' for c in normalized.lower())
    # Enlever les underscores multiples
    while '__' in normalized:
        normalized = normalized.replace('__', '_')
    # Enlever les underscores au début et à la fin
    return normalized.strip('_')


def get_column_value(row_dict, column_mapping, column_name):
    """Récupère la valeur d'une colonne en utilisant le mapping normalisé"""
    normalized = normalize_column_name(column_name)
    if normalized in column_mapping:
        mapped_name = column_mapping[normalized]
        return row_dict.get(mapped_name, "")
    return row_dict.get(column_name, "")


def normalize_segment_value(value):
    """Normalise une valeur de segment pour correspondre à SegmentEnum"""
    if not value:
        return "PARTICULIER"
    
    segment_mapping = {
        "PARTICULIER": "PARTICULIER",
        "PROFESSIONNEL": "PROFESSIONNEL", 
        "ENTREPRISE": "ENTREPRISE",
        "PERSONNEL": "PARTICULIER",
        "PERSONNE PHYSIQUE": "PARTICULIER",
        "MORALE": "ENTREPRISE",
        "PERSONNE MORALE": "ENTREPRISE",
        "PRO": "PROFESSIONNEL",
        "PROFESSION": "PROFESSIONNEL"
    }
    
    value_upper = str(value).upper().strip()
    return segment_mapping.get(value_upper, "PARTICULIER")


def normalize_produit_value(value):
    """Normalise une valeur de produit pour correspondre à ProduitEnum"""
    if not value:
        return "Conso"
    
    produit_mapping = {
        "CONSOMMATION": "Conso",
        "CONSO": "Conso",
        "HYPOTHECAIRE": "Hypothécaire",
        "IMMOBILIER": "Hypothécaire",
        "AUTOCAR": "Autocars",
        "TRÉSORERIE": "Trésorerie",
        "TRESORERIE": "Trésorerie",
        "PERSONNEL": "Conso",
        "PROFESSIONNEL": "Conso"
    }
    
    value_upper = str(value).upper().strip()
    return produit_mapping.get(value_upper, "Conso")


def normalize_statut_value(value):
    """Normalise une valeur de statut pour correspondre à StatutInterneEnum"""
    if not value:
        return "Impayé"
    
    statut_mapping = {
        "IMPAYÉ": "Impayé",
        "IMPAYE": "Impayé",
        "NORMAL": "Normal",
        "RESTRUCTURÉ": "Restructuré",
        "RESTRUCTURE": "Restructuré",
        "CONTENTIEUX": "Contentieux",
        "EN COURS": "Normal",
        "REGULIER": "Normal"
    }
    
    value_upper = str(value).upper().strip()
    return statut_mapping.get(value_upper, "Impayé")


def normalize_phone_number(phone: str) -> Optional[str]:
    """
    Normalise un numéro de téléphone au format 227XXXXXXXXX (sans le +)
    """
    if not phone:
        return None
    
    # Enlever tous les caractères non numériques
    phone_digits = ''.join(c for c in str(phone) if c.isdigit())
    
    # Si le numéro commence par 227 (indicatif du Niger), le garder
    if phone_digits.startswith('227'):
        return phone_digits
    # Sinon, ajouter 227 au début
    elif len(phone_digits) >= 8:
        return f"227{phone_digits[-8:]}"
    
    return None


def build_column_mapping(headers):
    """
    Construit un mapping intelligent entre les noms de colonnes Excel et les noms attendus
    Gère les variations de noms (tronqués, casse différente, etc.)
    """
    mapping = {}
    
    # Mapping exact avec variations possibles
    column_variations = {
        "dateSituation": ["dateSituation", "date_situation", "date situation", "datesituation", "date"],
        "refCredit": ["refCredit", "ref_credit", "ref credit", "reference", "ref", "credit"],
        "idClient": ["idClient", "id_client", "id client", "id", "clientid", "client_id"],
        "nomClient": ["nomClient", "nom_client", "nom client", "nom", "client", "clientname"],
        "telephoneClient": ["telephoneClient", "telephone_client", "telephone client", "telephone", "tel", "phone"],
        "segment": ["segment", "segmentation"],
        "agence": ["agence", "agency", "branch"],
        "gestionnaire": ["gestionnaire", "gestionnaire", "agent", "charge_client", "charge_client", "responsable", "manager"],
        "produit": ["produit", "product", "typeproduit"],
        "montantInitial": ["montantInitial", "montant_initial", "montant initial", "montant", "initial"],
        "encoursPrincipal": ["encoursPrincipal", "encours_principal", "encours principal", "encours", "principal"],
        "principalImpaye": ["principalImpaye", "principal_impaye", "principal impaye", "principal", "impaye"],
        "interetsImpayes": ["interetsImpayes", "interets_impayes", "interets impayes", "interets", "interest"],
        "penalitesImpayees": ["penalitesImpayees", "penalites_impayees", "penalites impayees", "penalites", "penalty"],
        "nbEcheancesImpayees": ["nbEcheancesImpayees", "nb_echeances_impayees", "nb echeances impayees", "echeances", "echeance"],
        "joursRetard": ["joursRetard", "jours_retard", "jours retard", "retard", "days"],
        "dateDerniereEcheanceImpayee": ["dateDerniereEcheanceImpayee", "date_derniere_echeance_impayee", "date derniere echeance impayee", "dateecheance"],
        "statutInterne": ["statutInterne", "statut_interne", "statut interne", "statut"],
        "garanties": ["garanties", "garantie", "guarantee"],
        "revenuMensuel": ["revenuMensuel", "revenu_mensuel", "revenu mensuel", "revenu", "salaire", "income"],
        "commentaire": ["commentaire", "comment", "remarque", "note"]
    }
    
    # Pour chaque en-tête normalisé, trouver la meilleure correspondance
    for normalized_target, variations in column_variations.items():
        best_match = None
        best_score = 0
        
        for header in headers:
            if not header:
                continue
                
            header_norm = normalize_column_name(header)
            
            # Correspondance exacte
            if header_norm == normalize_column_name(normalized_target):
                best_match = header
                best_score = 100
                break
            
            # Correspondance partielle
            for variation in variations:
                var_norm = normalize_column_name(variation)
                if header_norm == var_norm:
                    if len(variation) > best_score:
                        best_match = header
                        best_score = len(variation)
                        break
                elif var_norm in header_norm or header_norm in var_norm:
                    if len(var_norm) > best_score * 0.7:  # Seuil de 70%
                        best_match = header
                        best_score = len(var_norm) * 0.7
        
        if best_match:
            mapping[normalize_column_name(normalized_target)] = best_match
    
    return mapping


class ImpayesImportService:
    """Service unique pour le parsing des fichiers d'impayés"""
    
    @staticmethod
    async def parse_file_to_lignes(
        file: UploadFile,
        date_situation: str = None
    ) -> Tuple[List[LigneImpayeImport], Dict[str, Any]]:
        """
        Parse un fichier Excel/CSV et retourne les lignes + métadonnées
        
        Args:
            file: Fichier uploadé
            date_situation: Date de situation (optionnelle)
            
        Returns:
            Tuple[List[LigneImpayeImport], Dict]: Lignes parsées + métadonnées
        """
        try:
            contents = await file.read()
            filename = file.filename.lower()
            
            metadata = {
                "filename": file.filename,
                "size_bytes": len(contents),
                "extension": filename,
                "date_situation": date_situation
            }
            
            lignes = []
            
            if filename.endswith('.csv'):
                lignes, csv_metadata = await ImpayesImportService._parse_csv(
                    contents, filename, date_situation
                )
                metadata.update(csv_metadata)
            else:
                lignes, excel_metadata = await ImpayesImportService._parse_excel(
                    contents, filename, date_situation
                )
                metadata.update(excel_metadata)
            
            metadata["total_lignes"] = len(lignes)
            
            return lignes, metadata
            
        except Exception as e:
            raise Exception(f"Erreur lors du parsing du fichier: {str(e)}")
    
    @staticmethod
    async def _parse_csv(
        contents: bytes, 
        filename: str, 
        date_situation: str = None
    ) -> Tuple[List[LigneImpayeImport], Dict[str, Any]]:
        """Parse un fichier CSV"""
        try:
            # Gestion de l'encodage
            try:
                content_str = contents.decode('utf-8-sig')
            except UnicodeDecodeError:
                try:
                    content_str = contents.decode('latin-1')
                except UnicodeDecodeError:
                    content_str = contents.decode('cp1252')
            
            # Détection automatique du séparateur
            lines = content_str.split('\n') if '\n' in content_str else content_str.split('\r\n')
            first_line = lines[0] if lines else ""
            second_line = lines[1] if len(lines) > 1 else ""
            
            count_semicolon = first_line.count(';')
            count_comma = first_line.count(',')
            
            if second_line:
                count_semicolon += second_line.count(';')
                count_comma += second_line.count(',')
            
            delimiter = ';' if count_semicolon > count_comma else ','
            
            csv_reader = csv.DictReader(io.StringIO(content_str), delimiter=delimiter)
            csv_headers = csv_reader.fieldnames if csv_reader.fieldnames else []
            
            # Mapping intelligent des colonnes
            column_mapping = build_column_mapping(csv_headers)
            
            lignes = []
            for row_idx, row in enumerate(csv_reader, start=2):
                try:
                    ligne = ImpayesImportService._convert_row_to_ligne(
                        row, column_mapping, row_idx, date_situation
                    )
                    lignes.append(ligne)
                except Exception as e:
                    print(f"[ERROR] Erreur ligne {row_idx}: {str(e)}")
                    continue
            
            metadata = {
                "file_type": "csv",
                "delimiter": delimiter,
                "headers": csv_headers,
                "column_mapping": column_mapping
            }
            
            return lignes, metadata
            
        except Exception as e:
            raise Exception(f"Erreur lors du parsing CSV: {str(e)}")
    
    @staticmethod
    async def _parse_excel(
        contents: bytes, 
        filename: str, 
        date_situation: str = None
    ) -> Tuple[List[LigneImpayeImport], Dict[str, Any]]:
        """Parse un fichier Excel"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(io.BytesIO(contents), data_only=True)
            
            # Sélectionner la feuille
            if "Impayes" in wb.sheetnames:
                ws = wb["Impayes"]
            else:
                ws = wb.active
            
            # Lire les en-têtes
            headers = []
            for cell in ws[1]:
                try:
                    from app.routers.impayes import safe_cell_value
                    value = safe_cell_value(cell)
                    headers.append(value if value else "")
                except Exception:
                    headers.append("")
            
            # Mapping intelligent des colonnes
            column_mapping = build_column_mapping(headers)
            
            lignes = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Créer un dictionnaire avec les valeurs de la ligne
                    row_dict = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            if header:
                                from app.routers.impayes import safe_cell_value
                                row_dict[header] = safe_cell_value(cell)
                    
                    ligne = ImpayesImportService._convert_row_to_ligne(
                        row_dict, column_mapping, row_idx, date_situation
                    )
                    lignes.append(ligne)
                except Exception as e:
                    print(f"Erreur ligne {row_idx}: {str(e)}")
                    continue
            
            metadata = {
                "file_type": "excel",
                "sheet_name": ws.title,
                "headers": headers,
                "column_mapping": column_mapping
            }
            
            return lignes, metadata
            
        except ImportError:
            raise Exception("La bibliothèque openpyxl n'est pas installée. Veuillez l'installer avec: pip install openpyxl")
        except Exception as e:
            raise Exception(f"Erreur lors du parsing Excel: {str(e)}")
    
    @staticmethod
    def _convert_row_to_ligne(
        row: Dict[str, Any], 
        column_mapping: Dict[str, str], 
        row_idx: int,
        date_situation: str = None
    ) -> LigneImpayeImport:
        """Convert une ligne en objet LigneImpayeImport"""
        
        # Fonctions utilitaires de conversion
        def safe_float(value, default=0.0):
            try:
                if not value or value == '':
                    return default
                cleaned = str(value).strip().replace(' ', '').replace(',', '')
                return float(cleaned) if cleaned else default
            except (ValueError, TypeError):
                return default
        
        def safe_int(value, default=0):
            try:
                if not value or value == '':
                    return default
                cleaned = str(value).strip().replace(' ', '').replace(',', '')
                return int(float(cleaned)) if cleaned else default
            except (ValueError, TypeError):
                return default
        
        # Récupérer les valeurs avec le mapping intelligent
        agence_val = get_column_value(row, column_mapping, "agence") or ""
        produit_val = get_column_value(row, column_mapping, "produit") or ""
        
        # Construire refCredit si manquant
        ref_credit = get_column_value(row, column_mapping, "refCredit")
        if not ref_credit and agence_val and produit_val:
            produit_num = produit_val.split()[0] if produit_val.split() else ""
            ref_credit = f"{agence_val}{produit_num}" if produit_num else f"{agence_val}_{row_idx}"
        elif not ref_credit:
            ref_credit = f"CREDIT_{row_idx}"
        
        # Date de situation
        date_situation_val = get_column_value(row, column_mapping, "dateSituation")
        if not date_situation_val:
            date_situation_val = date_situation or ""
        
        # Créer l'objet LigneImpayeImport
        return LigneImpayeImport(
            dateSituation=date_situation_val,
            refCredit=ref_credit,
            idClient=get_column_value(row, column_mapping, "idClient") or f"CLIENT_{row_idx}",
            nomClient=get_column_value(row, column_mapping, "nomClient") or f"Client {row_idx}",
            telephoneClient=normalize_phone_number(get_column_value(row, column_mapping, "telephoneClient")),
            segment=SegmentEnum(normalize_segment_value(get_column_value(row, column_mapping, "segment"))),
            agence=agence_val,
            gestionnaire=get_column_value(row, column_mapping, "gestionnaire") or None,
            produit=ProduitEnum(normalize_produit_value(get_column_value(row, column_mapping, "produit"))),
            montantInitial=safe_float(get_column_value(row, column_mapping, "montantInitial"), 0),
            encoursPrincipal=safe_float(get_column_value(row, column_mapping, "encoursPrincipal"), 0),
            principalImpaye=safe_float(get_column_value(row, column_mapping, "principalImpaye"), 0),
            interetsImpayes=safe_float(get_column_value(row, column_mapping, "interetsImpayes"), 0),
            penalitesImpayees=safe_float(get_column_value(row, column_mapping, "penalitesImpayees"), 0),
            nbEcheancesImpayees=safe_int(get_column_value(row, column_mapping, "nbEcheancesImpayees"), 0),
            joursRetard=safe_int(get_column_value(row, column_mapping, "joursRetard"), 0),
            dateDerniereEcheanceImpayee=get_column_value(row, column_mapping, "dateDerniereEcheanceImpayee") or None,
            statutInterne=StatutInterneEnum(normalize_statut_value(get_column_value(row, column_mapping, "statutInterne"))),
            garanties=get_column_value(row, column_mapping, "garanties") or None,
            revenuMensuel=safe_float(get_column_value(row, column_mapping, "revenuMensuel")) if get_column_value(row, column_mapping, "revenuMensuel") else None,
            commentaire=get_column_value(row, column_mapping, "commentaire") or None,
        )
