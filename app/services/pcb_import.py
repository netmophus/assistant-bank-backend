"""
Service d'import Excel pour les comptes GL PCB UEMOA
"""
from datetime import datetime
from typing import Dict, Optional
from io import BytesIO
from app.models.pcb import create_or_update_gl_account


async def import_gl_from_excel(
    file_content: bytes,
    organization_id: str,
    date_solde: datetime
) -> Dict:
    """
    Importe les comptes GL depuis un fichier Excel
    
    IMPORTANT : Les données sont séparées par date. Chaque ligne peut avoir sa propre date
    (colonne Date_Solde). Si la colonne Date_Solde est vide, la date fournie lors de l'import
    sera utilisée. Les lignes avec le même Code_GL mais des dates différentes seront
    sauvegardées séparément.
    
    Format attendu (feuille GL_Import):
    - Code_GL: Code du compte (obligatoire)
    - Libelle_GL: Libellé (obligatoire)
    - Classe: Classe PCB UEMOA (1-7 pour bilan, 9 pour hors bilan. La classe 8 n'existe pas) (obligatoire)
    - Sous_classe: Optionnel
    - Solde_Debit: Montant débiteur
    - Solde_Credit: Montant créditeur
    - Solde_Net: Optionnel (prioritaire sur Solde_Debit/Solde_Credit)
    - Date_Solde: Date de solde (optionnel, formats acceptés: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY)
                   Si vide, utilise la date fournie lors de l'import
    - Devise: Optionnel (défaut XOF)
    
    Les lignes avec le même Code_GL seront agrégées (addition des débits et crédits)
    avant la sauvegarde, mais uniquement pour la même date_solde et devise.
    Exemple : Les données du 30/11/2025 et du 31/12/2025 seront séparées.
    """
    errors = []
    comptes_crees = 0
    comptes_mis_a_jour = 0
    
    try:
        # Importer openpyxl (gestion d'erreur si non installé)
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "total_lignes": 0,
                "comptes_crees": 0,
                "comptes_mis_a_jour": 0,
                "erreurs": [{"ligne": 0, "message": "La bibliothèque openpyxl n'est pas installée. Veuillez l'installer avec: pip install openpyxl"}]
            }
        
        # Lire le fichier Excel
        wb = load_workbook(BytesIO(file_content), data_only=True)
        
        # Chercher la feuille GL_Import
        if "GL_Import" not in wb.sheetnames:
            # Essayer la première feuille si GL_Import n'existe pas
            ws = wb.active
        else:
            ws = wb["GL_Import"]
        
        # Lire l'en-tête pour trouver les colonnes (insensible à la casse et aux espaces)
        headers = {}
        header_row = None
        normalized_headers = {}  # Pour la correspondance insensible à la casse
        
        # Mapping des noms de colonnes possibles (normalisés)
        column_mapping = {
            "code_gl": "Code_GL",
            "codegl": "Code_GL",
            "code": "Code_GL",
            "libelle_gl": "Libelle_GL",
            "libellegl": "Libelle_GL",
            "libelle": "Libelle_GL",
            "libellé": "Libelle_GL",
            "classe": "Classe",
            "sous_classe": "Sous_classe",
            "sousclasse": "Sous_classe",
            "solde_debit": "Solde_Debit",
            "soldedebit": "Solde_Debit",
            "solde_débit": "Solde_Debit",
            "solde_credit": "Solde_Credit",
            "soldecredit": "Solde_Credit",
            "solde_crédit": "Solde_Credit",
            "solde_net": "Solde_Net",
            "soldenet": "Solde_Net",
            "date_solde": "Date_Solde",
            "datesolde": "Date_Solde",
            "devise": "Devise",
            "type": "Type",
        }
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            row_values = [str(cell.value).strip().lower().replace(" ", "_").replace("é", "e") if cell.value else "" for cell in row]
            
            # Vérifier si cette ligne contient "code_gl" ou variantes
            if any("code" in val and "gl" in val for val in row_values) or any("code_gl" in val for val in row_values):
                header_row = row_idx
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value:
                        original_name = str(cell.value).strip()
                        normalized_name = original_name.lower().replace(" ", "_").replace("é", "e").replace("_", "")
                        
                        # Chercher dans le mapping
                        matched_name = None
                        for key, value in column_mapping.items():
                            if normalized_name == key or normalized_name.startswith(key) or key in normalized_name:
                                matched_name = value
                                break
                        
                        # Utiliser le nom normalisé ou l'original
                        final_name = matched_name if matched_name else original_name
                        headers[final_name] = col_idx
                        normalized_headers[normalized_name] = final_name
                break
        
        if not header_row:
            # Essayer de trouver toutes les colonnes disponibles pour le message d'erreur
            available_cols = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=False), start=1):
                for cell in row:
                    if cell.value:
                        available_cols.append(str(cell.value).strip())
                if available_cols:
                    break
            
            return {
                "total_lignes": 0,
                "comptes_crees": 0,
                "comptes_mis_a_jour": 0,
                "erreurs": [{
                    "ligne": 0,
                    "message": f"En-tête 'Code_GL' introuvable dans les 10 premières lignes. "
                              f"Colonnes trouvées: {', '.join(available_cols[:10]) if available_cols else 'Aucune'}. "
                              f"Vérifiez que la première ligne contient les colonnes: Code_GL, Libelle_GL, Classe"
                }]
            }
        
        # Vérifier les colonnes obligatoires
        required_columns = ["Code_GL", "Libelle_GL", "Classe"]
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            found_cols = list(headers.keys())
            return {
                "total_lignes": 0,
                "comptes_crees": 0,
                "comptes_mis_a_jour": 0,
                "erreurs": [{
                    "ligne": header_row,
                    "message": f"Colonnes obligatoires manquantes: {', '.join(missing_columns)}. "
                              f"Colonnes trouvées: {', '.join(found_cols)}"
                }]
            }
        
        # Fonction helper pour récupérer une valeur de cellule
        def get_cell_value(row, col_name):
            col_idx = headers.get(col_name)
            if col_idx and col_idx <= len(row):
                cell = row[col_idx - 1]
                return cell.value if cell.value is not None else ""
            return ""
        
        # Dictionnaire pour regrouper les lignes par Code_GL + Date_Solde + Devise
        # Clé : (code_gl, date_solde, devise)
        gl_aggregated = {}
        
        # Traiter chaque ligne et agréger
        total_lignes = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            total_lignes += 1
            index = row_idx - header_row - 1
            try:
                
                # Récupérer les valeurs
                code_gl = str(get_cell_value(row, "Code_GL") or "").strip()
                libelle_gl = str(get_cell_value(row, "Libelle_GL") or "").strip()
                classe = get_cell_value(row, "Classe")
                
                # Validation de base
                if not code_gl:
                    errors.append({"ligne": index + 2, "code": code_gl, "message": "Code GL vide"})
                    continue
                
                if not libelle_gl:
                    errors.append({"ligne": index + 2, "code": code_gl, "message": "Libellé vide"})
                    continue
                
                try:
                    classe = int(classe)
                    # Classes valides : 1-7 (bilan) et 9 (hors bilan). La classe 8 n'existe pas dans le PCB UEMOA
                    valid_classes = [1, 2, 3, 4, 5, 6, 7, 9]
                    if classe not in valid_classes:
                        errors.append({"ligne": index + 2, "code": code_gl, "message": f"Classe invalide: {classe} (doit être 1-7 pour bilan ou 9 pour hors bilan. La classe 8 n'existe pas)"})
                        continue
                except (ValueError, TypeError):
                    errors.append({"ligne": index + 2, "code": code_gl, "message": f"Classe invalide: {classe}"})
                    continue
                
                # Récupérer les soldes
                solde_debit_val = get_cell_value(row, "Solde_Debit")
                solde_credit_val = get_cell_value(row, "Solde_Credit")
                solde_net_val = get_cell_value(row, "Solde_Net")
                
                try:
                    solde_debit = float(solde_debit_val) if solde_debit_val else 0.0
                except (ValueError, TypeError):
                    solde_debit = 0.0
                
                try:
                    solde_credit = float(solde_credit_val) if solde_credit_val else 0.0
                except (ValueError, TypeError):
                    solde_credit = 0.0
                
                try:
                    solde_net = float(solde_net_val) if solde_net_val else None
                except (ValueError, TypeError):
                    solde_net = None
                
                # Date de solde : utiliser celle du fichier Excel si présente, sinon celle fournie lors de l'import
                # Cela permet d'importer des fichiers avec plusieurs dates (ex: plusieurs mois)
                # Les données seront séparées par date grâce à la clé d'agrégation (code_gl, date_solde, devise)
                # IMPORTANT : Normaliser toutes les dates à minuit (00:00:00) pour éviter les problèmes de comparaison
                date_solde_row = get_cell_value(row, "Date_Solde")
                if date_solde_row:
                    if isinstance(date_solde_row, datetime):
                        # Normaliser à minuit pour éviter les problèmes de comparaison
                        date_solde_final = datetime(date_solde_row.year, date_solde_row.month, date_solde_row.day)
                    elif isinstance(date_solde_row, str):
                        try:
                            # Essayer différents formats de date
                            date_parsed = datetime.strptime(date_solde_row.strip(), "%Y-%m-%d")
                            date_solde_final = datetime(date_parsed.year, date_parsed.month, date_parsed.day)
                        except:
                            try:
                                date_parsed = datetime.strptime(date_solde_row.strip(), "%d/%m/%Y")
                                date_solde_final = datetime(date_parsed.year, date_parsed.month, date_parsed.day)
                            except:
                                try:
                                    date_parsed = datetime.strptime(date_solde_row.strip(), "%d-%m-%Y")
                                    date_solde_final = datetime(date_parsed.year, date_parsed.month, date_parsed.day)
                                except:
                                    # Si aucun format ne fonctionne, utiliser la date fournie lors de l'import
                                    date_solde_final = datetime(date_solde.year, date_solde.month, date_solde.day)
                    else:
                        date_solde_final = datetime(date_solde.year, date_solde.month, date_solde.day)
                else:
                    # Si aucune date dans le fichier, utiliser celle fournie lors de l'import (normalisée à minuit)
                    date_solde_final = datetime(date_solde.year, date_solde.month, date_solde.day)
                
                # Devise
                devise_val = get_cell_value(row, "Devise")
                devise = str(devise_val).strip() if devise_val else "XOF"
                if not devise:
                    devise = "XOF"
                
                # Type (déduire de la classe si non fourni)
                type_compte_val = get_cell_value(row, "Type")
                type_compte = str(type_compte_val).strip() if type_compte_val else ""
                if not type_compte:
                    if classe in [1, 2, 3, 4, 5]:
                        # Bilan
                        if classe in [1, 2]:
                            type_compte = "passif"
                        else:
                            type_compte = "actif"
                    elif classe == 6:
                        type_compte = "charge"
                    elif classe == 7:
                        type_compte = "produit"
                    elif classe == 9:
                        type_compte = "hors_bilan"  # Hors bilan
                
                # Préparer les données
                sous_classe_val = get_cell_value(row, "Sous_classe")
                sous_classe = str(sous_classe_val).strip() if sous_classe_val else None
                
                # Clé d'agrégation : (code_gl, date_solde, devise)
                aggregation_key = (code_gl, date_solde_final, devise)
                
                # Agréger les données
                if aggregation_key not in gl_aggregated:
                    gl_aggregated[aggregation_key] = {
                        "code": code_gl,
                        "libelle": libelle_gl,  # Prendre le libellé de la première ligne
                        "classe": classe,
                        "sous_classe": sous_classe,  # Prendre la sous_classe de la première ligne
                        "type": type_compte,
                        "nature": "compte_detail",
                        "solde_debit": solde_debit,
                        "solde_credit": solde_credit,
                        "solde_net": solde_net,  # Si solde_net est fourni, on garde le dernier
                        "date_solde": date_solde_final,
                        "devise": devise,
                        "is_active": True,
                        "nb_lignes": 1,  # Compteur pour debug
                    }
                else:
                    # Additionner les soldes pour le même Code_GL
                    gl_aggregated[aggregation_key]["solde_debit"] += solde_debit
                    gl_aggregated[aggregation_key]["solde_credit"] += solde_credit
                    gl_aggregated[aggregation_key]["nb_lignes"] += 1
                    # Si solde_net est fourni dans cette ligne, on le met à jour
                    if solde_net is not None:
                        gl_aggregated[aggregation_key]["solde_net"] = solde_net
                    
            except Exception as e:
                try:
                    code_gl_error = get_cell_value(row, "Code_GL") or "N/A"
                except:
                    code_gl_error = "N/A"
                errors.append({
                    "ligne": row_idx,
                    "code": str(code_gl_error),
                    "message": f"Erreur lors du traitement: {str(e)}"
                })
        
        # Maintenant, sauvegarder les données agrégées
        from app.models.pcb import get_gl_account_by_code
        
        for aggregation_key, gl_data in gl_aggregated.items():
            try:
                # Retirer le compteur nb_lignes avant la sauvegarde
                nb_lignes = gl_data.pop("nb_lignes", 1)
                
                # Vérifier si le compte existe déjà
                existing = await get_gl_account_by_code(gl_data["code"], organization_id, gl_data["date_solde"])
                
                # Créer ou mettre à jour
                await create_or_update_gl_account(gl_data, organization_id)
                
                if existing:
                    comptes_mis_a_jour += 1
                else:
                    comptes_crees += 1
                    
            except Exception as e:
                errors.append({
                    "ligne": 0,
                    "code": gl_data.get("code", "N/A"),
                    "message": f"Erreur lors de la sauvegarde du compte agrégé: {str(e)}"
                })
        
        return {
            "total_lignes": total_lignes,
            "comptes_crees": comptes_crees,
            "comptes_mis_a_jour": comptes_mis_a_jour,
            "erreurs": errors
        }
        
    except Exception as e:
        return {
            "total_lignes": 0,
            "comptes_crees": 0,
            "comptes_mis_a_jour": 0,
            "erreurs": [{"ligne": 0, "message": f"Erreur lors de la lecture du fichier: {str(e)}"}]
        }

