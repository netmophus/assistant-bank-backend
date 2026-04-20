"""
Génération automatique du bilan + compte de résultat à partir d'une balance GL.

Entrée : fichier Excel avec colonnes
    Code_GL, Libelle_GL, Classe, Solde_Debit, Solde_Credit, Date_Solde, Devise

Sortie : structure {bilan_actif, bilan_passif, compte_resultat, totaux, meta}
avec soldes exprimés en millions de FCFA.

Principe :
1. Parse Excel → liste brute de lignes GL
2. Agrégation par Code_GL (somme Débit/Crédit sur toutes les agences/lignes)
3. Mapping par préfixe du code vers un poste réglementaire PCB UMOA
4. Règle de signe :
    - Actif / Charges  : solde = Débit − Crédit
    - Passif / Produits: solde = Crédit − Débit
5. Agrégation des soldes GL par poste de destination
6. Calcul des totaux (total actif, total passif, produits, charges, résultat net)

Conversion d'unité : les montants Excel sont en XOF bruts, on restitue en M FCFA
(division par 1_000_000) pour cohérence avec le reste du module PCB.
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None  # type: ignore


# ─────────────────────────────────────────────────────────────────────────────
# Mapping officiel PCB UMOA : préfixe GL → poste réglementaire
# Les préfixes sont testés par longueur décroissante (longest-match wins), de
# sorte que '301' est détecté avant '30'.
# ─────────────────────────────────────────────────────────────────────────────

# Format : (prefix, code_poste, libelle, bucket)
# bucket ∈ {'actif', 'passif', 'charges', 'produits'}
PREFIX_MAPPING: List[Tuple[str, str, str, str]] = [
    # ─── BILAN ACTIF (BA1 → BA14) ──────────────────────────────
    ('10',  'BA1',  'Caisse et correspondants BCEAO',               'actif'),
    ('11',  'BA1',  'Caisse et correspondants BCEAO',               'actif'),
    ('301', 'BA2',  'Créances sur établissements bancaires',        'actif'),
    ('311', 'BA2',  'Créances sur établissements bancaires',        'actif'),
    ('321', 'BA2',  'Créances sur établissements bancaires',        'actif'),
    ('12',  'BA3',  'Crédits à la clientèle — court terme',         'actif'),
    ('13',  'BA3',  'Crédits à la clientèle — moyen/long terme',    'actif'),
    ('20',  'BA4',  'Titres de placement',                          'actif'),
    ('21',  'BA4',  'Titres de placement',                          'actif'),
    ('22',  'BA5',  "Titres d'investissement",                      'actif'),
    ('23',  'BA5',  "Titres d'investissement",                      'actif'),
    ('24',  'BA6',  'Titres de participation',                      'actif'),
    ('28',  'BA9',  'Autres actifs',                                'actif'),
    ('29',  'BA10', 'Comptes de régularisation actif',              'actif'),
    ('40',  'BA11', 'Immobilisations incorporelles',                'actif'),
    ('41',  'BA11', 'Immobilisations incorporelles',                'actif'),
    ('42',  'BA12', 'Immobilisations corporelles',                  'actif'),
    ('43',  'BA12', 'Immobilisations corporelles',                  'actif'),

    # ─── BILAN PASSIF (BP1 → BP9) ──────────────────────────────
    ('15',  'BP1',  'Dettes envers les établissements bancaires',   'passif'),
    ('16',  'BP2',  'Dettes à terme',                               'passif'),
    ('17',  'BP2',  'Dettes à terme',                               'passif'),
    ('25',  'BP3',  'Dépôts de la clientèle à vue',                 'passif'),
    ('26',  'BP3',  'Dépôts de la clientèle à terme',               'passif'),
    ('27',  'BP3',  "Dépôts et comptes d'épargne",                  'passif'),
    ('33',  'BP4',  'Autres passifs',                               'passif'),
    ('34',  'BP4',  'Autres passifs',                               'passif'),
    ('35',  'BP5',  'Comptes de régularisation passif',             'passif'),
    ('36',  'BP5',  'Comptes de régularisation passif',             'passif'),
    ('38',  'BP6',  'Provisions pour risques et charges',           'passif'),
    ('51',  'BP7',  'Capital social',                               'passif'),
    ('53',  'BP7',  'Réserves et primes',                           'passif'),
    ('55',  'BP8',  "Résultat en instance d'affectation",           'passif'),
    ('56',  'BP8',  "Résultat en instance d'affectation",           'passif'),
    ('57',  'BP8',  "Résultat en instance d'affectation",           'passif'),
    ('58',  'BP8',  "Résultat en instance d'affectation",           'passif'),
    ('59',  'BP9',  "Résultat de l'exercice",                       'passif'),

    # ─── COMPTE DE RÉSULTAT (R1 → R12) ─────────────────────────
    ('70',  'R1',   'Intérêts et produits assimilés',               'produits'),
    ('60',  'R2',   'Intérêts et charges assimilées',               'charges'),
    ('71',  'R3',   'Produits sur crédit-bail et assimilés',        'produits'),
    ('61',  'R4',   'Charges sur crédit-bail et assimilées',        'charges'),
    ('72',  'R5',   'Commissions (produits)',                       'produits'),
    ('73',  'R6',   'Gains sur opérations financières',             'produits'),
    ('74',  'R6',   'Gains sur opérations financières',             'produits'),
    ('75',  'R6',   "Autres produits d'exploitation bancaire",      'produits'),
    ('63',  'R7',   "Autres charges d'exploitation bancaire",       'charges'),
    ('64',  'R7',   "Autres charges d'exploitation bancaire",       'charges'),
    ('65',  'R7',   "Autres charges d'exploitation bancaire",       'charges'),
    ('66',  'R8',   "Charges générales d'exploitation",             'charges'),
    ('67',  'R9',   'Dotations aux amortissements et provisions',   'charges'),
    ('68',  'R9',   'Dotations aux amortissements et provisions',   'charges'),
    ('69',  'R10',  'Impôts sur les bénéfices',                     'charges'),
    ('76',  'R11',  'Produits exceptionnels',                       'produits'),
    ('77',  'R11',  'Produits exceptionnels',                       'produits'),
    ('78',  'R12',  'Reprises de provisions',                       'produits'),
]

# Tri par longueur de préfixe décroissante pour garantir le longest-match
PREFIX_MAPPING_SORTED = sorted(PREFIX_MAPPING, key=lambda e: len(e[0]), reverse=True)

# Libellé canonique par code (première occurrence du mapping fait foi)
POSTE_LIBELLES: Dict[str, str] = {}
for _prefix, _code, _lib, _bucket in PREFIX_MAPPING:
    POSTE_LIBELLES.setdefault(_code, _lib)


def _find_poste(code_gl: str) -> Optional[Tuple[str, str, str]]:
    """
    Retourne (code_poste, libelle, bucket) pour un code GL, ou None si aucun
    préfixe ne correspond. Utilise le longest-match.
    """
    if not code_gl:
        return None
    for prefix, code, libelle, bucket in PREFIX_MAPPING_SORTED:
        if code_gl.startswith(prefix):
            return code, libelle, bucket
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Parsing Excel
# ─────────────────────────────────────────────────────────────────────────────

# Synonymes tolérés pour les en-têtes de colonnes
COLUMN_ALIASES = {
    'code_gl': ['code_gl', 'code gl', 'compte', 'compte_gl', 'numero_compte', 'code', 'n_compte'],
    'libelle_gl': ['libelle_gl', 'libelle gl', 'libelle', 'libellé', 'intitule', 'intitulé'],
    'classe': ['classe', 'class'],
    'solde_debit': ['solde_debit', 'solde_débit', 'solde debit', 'solde débit', 'debit', 'débit'],
    'solde_credit': ['solde_credit', 'solde_crédit', 'solde credit', 'solde crédit', 'credit', 'crédit'],
    'date_solde': ['date_solde', 'date solde', 'date'],
    'devise': ['devise', 'currency'],
}


def _norm_header(s) -> str:
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip().lower())


def _parse_numeric(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        f = float(v)
        return 0.0 if f != f else f  # exclut NaN
    if isinstance(v, str):
        s = v.strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
        if not s:
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def _detect_columns(rows_preview: List[tuple]) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    Cherche dans les premières lignes celle qui ressemble à un en-tête et
    retourne (row_index, {col_key: col_idx}).
    """
    for r_idx, row in enumerate(rows_preview):
        if not row:
            continue
        headers = [_norm_header(c) for c in row]
        found: Dict[str, int] = {}
        for key, aliases in COLUMN_ALIASES.items():
            for c_idx, h in enumerate(headers):
                if h in aliases:
                    found[key] = c_idx
                    break
        # Il faut au minimum code_gl + solde_debit + solde_credit
        if 'code_gl' in found and 'solde_debit' in found and 'solde_credit' in found:
            return r_idx, found
    return None


def parse_gl_excel(file_bytes: bytes) -> List[Dict]:
    """
    Parse le fichier Excel de balance GL et retourne une liste brute :
        [{code_gl, libelle_gl, classe, debit, credit, date_solde, devise}, ...]
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl n'est pas installé. Exécutez `pip install openpyxl`.")

    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    for ws in wb.worksheets:
        # Lit les 20 premières lignes pour détecter l'en-tête
        preview = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
        detected = _detect_columns(preview)
        if not detected:
            continue

        header_row_idx, cols = detected
        rows: List[Dict] = []

        # Parcourt les lignes après l'en-tête
        for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
            if not row or all(c is None for c in row):
                continue

            def get(key: str):
                idx = cols.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            code_gl = get('code_gl')
            if code_gl is None:
                continue
            code_str = str(code_gl).strip()
            # Ignore les lignes de total / sous-total
            if not code_str or not re.match(r'^\d', code_str):
                continue

            rows.append({
                'code_gl': code_str,
                'libelle_gl': (str(get('libelle_gl') or '').strip() or code_str),
                'classe': str(get('classe') or '').strip(),
                'debit': _parse_numeric(get('solde_debit')),
                'credit': _parse_numeric(get('solde_credit')),
                'date_solde': get('date_solde'),
                'devise': str(get('devise') or '').strip() or 'XOF',
            })

        if rows:
            return rows

    # Aucune feuille exploitable : diagnostic
    sheet_names = ", ".join(ws.title for ws in wb.worksheets)
    raise ValueError(
        f"Aucun en-tête de balance GL reconnu dans le fichier. "
        f"Feuilles examinées : {sheet_names}. "
        f"Colonnes attendues : Code_GL, Libelle_GL, Classe, Solde_Debit, "
        f"Solde_Credit, Date_Solde, Devise."
    )


# ─────────────────────────────────────────────────────────────────────────────
# Agrégation + mapping
# ─────────────────────────────────────────────────────────────────────────────

def _aggregate_by_code(rows: List[Dict]) -> Dict[str, Dict]:
    """Regroupe les lignes par code_gl (somme débit/crédit)."""
    agg: Dict[str, Dict] = {}
    for r in rows:
        code = r['code_gl']
        cur = agg.setdefault(code, {
            'code_gl': code,
            'libelle_gl': r['libelle_gl'],
            'debit': 0.0,
            'credit': 0.0,
            'nb_rows': 0,
        })
        cur['debit'] += r['debit']
        cur['credit'] += r['credit']
        cur['nb_rows'] += 1
    return agg


def _compute_solde(bucket: str, debit: float, credit: float) -> float:
    """Règle de signe PCB UMOA."""
    if bucket in ('actif', 'charges'):
        return debit - credit
    # passif, produits
    return credit - debit


def generer_bilan_et_cr(
    file_bytes: bytes,
    *,
    date_cloture: str,
    date_n1: Optional[str] = None,
    date_realisation: Optional[str] = None,
) -> Dict:
    """
    Point d'entrée principal : lit l'Excel, applique le mapping, retourne la
    structure attendue par le frontend (soldes en millions de FCFA).
    """
    rows = parse_gl_excel(file_bytes)
    if not rows:
        raise ValueError("Le fichier ne contient aucune ligne GL exploitable.")

    by_code = _aggregate_by_code(rows)

    # Regroupement par poste de destination
    #   postes[code_poste] = {code, libelle, solde_bruts, nb_gl, gl_codes, bucket}
    postes: Dict[str, Dict] = {}
    non_mappes: List[Dict] = []

    for code_gl, info in by_code.items():
        match = _find_poste(code_gl)
        if not match:
            solde = info['debit'] - info['credit']
            non_mappes.append({
                'code_gl': code_gl,
                'libelle_gl': info['libelle_gl'],
                'solde_brut': solde,
            })
            continue

        code_poste, libelle, bucket = match
        solde_gl = _compute_solde(bucket, info['debit'], info['credit'])

        p = postes.setdefault(code_poste, {
            'code': code_poste,
            'libelle': POSTE_LIBELLES.get(code_poste, libelle),
            'bucket': bucket,
            'solde_bruts': 0.0,
            'nb_gl': 0,
        })
        p['solde_bruts'] += solde_gl
        p['nb_gl'] += 1

    # Sérialisation en millions de FCFA
    def _to_millions(x: float) -> float:
        return round(x / 1_000_000.0, 2)

    def _serialize(p: Dict) -> Dict:
        return {
            'code': p['code'],
            'libelle': p['libelle'],
            'solde': _to_millions(p['solde_bruts']),
            'nb_gl': p['nb_gl'],
        }

    # Split par bucket + tri par code (ordre BA1, BA2, …)
    def _code_sort_key(code: str) -> Tuple[str, int]:
        m = re.match(r'^([A-Z]+)(\d+)$', code)
        if m:
            return (m.group(1), int(m.group(2)))
        return (code, 0)

    actif_postes = sorted(
        [p for p in postes.values() if p['bucket'] == 'actif'],
        key=lambda p: _code_sort_key(p['code']),
    )
    passif_postes = sorted(
        [p for p in postes.values() if p['bucket'] == 'passif'],
        key=lambda p: _code_sort_key(p['code']),
    )
    cr_postes = sorted(
        [p for p in postes.values() if p['bucket'] in ('charges', 'produits')],
        key=lambda p: _code_sort_key(p['code']),
    )

    total_actif_bruts = sum(p['solde_bruts'] for p in actif_postes)
    total_passif_bruts = sum(p['solde_bruts'] for p in passif_postes)
    total_produits_bruts = sum(p['solde_bruts'] for p in postes.values() if p['bucket'] == 'produits')
    total_charges_bruts = sum(p['solde_bruts'] for p in postes.values() if p['bucket'] == 'charges')
    resultat_net_bruts = total_produits_bruts - total_charges_bruts

    return {
        'bilan_actif': [_serialize(p) for p in actif_postes],
        'bilan_passif': [_serialize(p) for p in passif_postes],
        'compte_resultat': [_serialize(p) for p in cr_postes],
        'totaux': {
            'total_actif': _to_millions(total_actif_bruts),
            'total_passif': _to_millions(total_passif_bruts),
            'total_produits': _to_millions(total_produits_bruts),
            'total_charges': _to_millions(total_charges_bruts),
            'resultat_net': _to_millions(resultat_net_bruts),
        },
        'meta': {
            'date_cloture': date_cloture,
            'date_n1': date_n1,
            'date_realisation': date_realisation,
            'nb_lignes_gl': len(rows),
            'nb_gl_distincts': len(by_code),
            'nb_gl_non_mappes': len(non_mappes),
            'gl_non_mappes': non_mappes[:50],  # aperçu
        },
    }
