"""
Endpoints pour la configuration des impayés.
"""
from fastapi import APIRouter, HTTPException, status, Depends
from fastapi.responses import StreamingResponse
import os

from app.schemas.impayes_config import (
    ImpayesConfig,
    ImpayesConfigPublic,
    TrancheRetard,
    ModeleSMS,
    RegleRestructuration,
    ParametresTechniques,
)
from app.models.impayes_config import (
    get_impayes_config,
    create_or_update_impayes_config,
)
from app.core.deps import get_current_user

router = APIRouter(
    prefix="/impayes/config",
    tags=["impayes-config"],
)


@router.get("/tranches")
async def get_tranches_config_endpoint(
    current_user: dict = Depends(get_current_user),
):
    """Récupère uniquement les tranches de retard pour l'organisation (accessible aux utilisateurs)"""
    org_id = str(current_user.get("organization_id"))
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Vous devez être rattaché à une organisation.",
        )
    
    config = await get_impayes_config(org_id)
    if not config:
        return {"tranches_retard": []}
    
    return {"tranches_retard": config.get("tranches_retard", [])}


@router.get("", response_model=ImpayesConfigPublic)
async def get_impayes_config_endpoint(
    current_user: dict = Depends(get_current_user),
):
    """Récupère la configuration des impayés pour l'organisation"""
    org_id = str(current_user.get("organization_id"))
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Vous devez être rattaché à une organisation.",
        )
    
    config = await get_impayes_config(org_id)
    if not config:
        # Créer une config par défaut
        default_config = ImpayesConfig(organization_id=org_id)
        config = await create_or_update_impayes_config(org_id, default_config)
    
    return config


@router.put("", response_model=ImpayesConfigPublic)
async def update_impayes_config_endpoint(
    config: ImpayesConfig,
    current_user: dict = Depends(get_current_user),
):
    """Met à jour la configuration des impayés"""
    org_id = str(current_user.get("organization_id"))
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Vous devez être rattaché à une organisation.",
        )
    
    # S'assurer que l'organization_id correspond
    config.organization_id = org_id
    
    updated_config = await create_or_update_impayes_config(org_id, config)
    return updated_config


@router.post("/init-modeles-sms")
async def init_default_sms_models(
    current_user: dict = Depends(get_current_user),
):
    """Initialise les modèles SMS par défaut pour l'organisation"""
    from app.schemas.impayes_config import _get_default_modeles_sms
    
    org_id = str(current_user.get("organization_id"))
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Vous devez être rattaché à une organisation.",
        )
    
    config = await get_impayes_config(org_id)
    if not config:
        # Créer une config par défaut
        default_config = ImpayesConfig(organization_id=org_id)
        config = await create_or_update_impayes_config(org_id, default_config)
    
    # Récupérer les modèles par défaut
    default_modeles = _get_default_modeles_sms()
    modeles_dict = [m.model_dump() for m in default_modeles]
    
    # Mettre à jour la config
    config_obj = ImpayesConfig(
        organization_id=org_id,
        tranches_retard=config.get("tranches_retard", []),
        regle_restructuration=config.get("regle_restructuration", {}),
        modeles_sms=default_modeles,
        parametres_techniques=config.get("parametres_techniques", {})
    )
    
    updated_config = await create_or_update_impayes_config(org_id, config_obj)
    
    return {
        "message": f"{len(default_modeles)} modèles SMS par défaut initialisés avec succès",
        "config": updated_config
    }


@router.get("/modele-excel")
async def download_excel_template():
    """Télécharge le modèle Excel pour l'import des impayés"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        # Créer un workbook en mémoire
        wb = Workbook()
        ws = wb.active
        ws.title = "Impayes"
        
        # En-têtes avec style
        headers = [
            "dateSituation",
            "refCredit",
            "idClient",
            "nomClient",
            "telephoneClient",
            "segment",
            "agence",
            "gestionnaire",
            "produit",
            "montantInitial",
            "encoursPrincipal",
            "principalImpayé",
            "interetsImpayés",
            "penalitesImpayées",
            "nbEcheancesImpayées",
            "joursRetard",
            "dateDerniereEcheanceImpayee",
            "statutInterne",
            "garanties",
            "revenuMensuel",
            "commentaire"
        ]
        
        # Style pour les en-têtes
        header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Ajouter les en-têtes
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        
        # Ajouter une ligne d'exemple (optionnel)
        example_row = [
            "2025-01-15",
            "CRED-2024-001",
            "CLI-12345",
            "kane lawan",
            "22796648383",
            "PARTICULIER",
            "AG001",
            "Ali Diallo",
            "Conso",
            "5000000",
            "3000000",
            "500000",
            "75000",
            "25000",
            "2",
            "45",
            "2024-12-01",
            "Impayé",
            "Hypothèque",
            "250000",
            "Client à contacter"
        ]
        
        for col, value in enumerate(example_row, start=1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Sauvegarder dans un buffer en mémoire
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="modele_impayes.xlsx"'
            },
        )
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="La bibliothèque openpyxl n'est pas installée. Veuillez l'installer avec: pip install openpyxl",
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la génération du modèle Excel: {str(e)}",
        )

