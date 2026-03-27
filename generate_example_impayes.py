"""
Script pour générer un fichier Excel d'exemple avec des données fictives pour l'import des impayés
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random

# Créer un nouveau workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Impayes"

# En-têtes
headers = [
    "dateSituation",
    "refCredit",
    "idClient",
    "nomClient",
    "telephoneClient",
    "segment",
    "agence",
    "produit",
    "montantInitial",
    "encoursPrincipal",
    "principalImpayé",
    "interetsImpayés",
    "penalitesImpayées",
    "nbEcheancesImpayées",
    "joursRetard",
    "dateDerniereEcheanceImpayee",
    "statutInterne",
    "garanties",
    "revenuMensuel",
    "commentaire"
]

# Style pour les en-têtes
header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Ajouter les en-têtes
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Ajuster la largeur des colonnes
for col in range(1, len(headers) + 1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

# Données fictives
segments = ["PARTICULIER", "PME", "PMI"]
agences = ["AG001", "AG002", "AG003", "AG004"]
produits = ["Conso", "Immo", "Trésorerie", "Autre"]
statuts = ["Normal", "Impayé", "Douteux", "Compromis"]
garanties = ["Hypothèque", "Caution", "Nantissement", "Gage", "Sans garantie"]
noms = [
    "DUPONT Jean", "MARTIN Marie", "BERNARD Pierre", "DUBOIS Sophie",
    "LAURENT Michel", "SIMON Claire", "MICHEL Anne", "GARCIA Carlos",
    "RODRIGUEZ Maria", "WILSON John", "BROWN Sarah", "DAVIS Robert",
    "MILLER Jennifer", "MOORE David", "TAYLOR Lisa", "ANDERSON Mark",
    "THOMAS Patricia", "JACKSON Daniel", "WHITE Susan", "HARRIS Christopher"
]

# Générer 20 lignes de données fictives
date_situation = datetime.now().strftime("%Y-%m-%d")
base_date = datetime.now()

for row in range(2, 22):  # 20 lignes de données
    jours_retard = random.randint(1, 120)
    date_echeance = (base_date - timedelta(days=jours_retard)).strftime("%Y-%m-%d")
    
    segment = random.choice(segments)
    montant_initial = random.randint(1000000, 50000000)  # 1M à 50M FCFA
    encours_principal = random.randint(500000, int(montant_initial * 0.8))
    principal_impaye = random.randint(50000, int(encours_principal * 0.3))
    interets_impayes = random.randint(5000, int(principal_impaye * 0.2))
    penalites = random.randint(1000, int(principal_impaye * 0.1))
    nb_echeances = random.randint(1, 6)
    revenu_mensuel = random.randint(100000, 2000000) if segment == "PARTICULIER" else random.randint(500000, 10000000)
    
    data = [
        date_situation,  # dateSituation
        f"CRED-2024-{row-1:04d}",  # refCredit
        f"CLI-{random.randint(10000, 99999)}",  # idClient
        random.choice(noms),  # nomClient
        f"+227{random.randint(90000000, 99999999)}",  # telephoneClient
        segment,  # segment
        random.choice(agences),  # agence
        random.choice(produits),  # produit
        montant_initial,  # montantInitial
        encours_principal,  # encoursPrincipal
        principal_impaye,  # principalImpayé
        interets_impayes,  # interetsImpayés
        penalites,  # penalitesImpayées
        nb_echeances,  # nbEcheancesImpayées
        jours_retard,  # joursRetard
        date_echeance,  # dateDerniereEcheanceImpayee
        random.choice(statuts),  # statutInterne
        random.choice(garanties),  # garanties
        revenu_mensuel,  # revenuMensuel
        random.choice([
            "Client à contacter",
            "Relance en cours",
            "Négociation en cours",
            "Dossier en restructuration",
            "Procédure judiciaire",
            "Aucun commentaire"
        ])  # commentaire
    ]
    
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Sauvegarder le fichier
filename = "exemple_impayes.xlsx"
wb.save(filename)
print(f"✅ Fichier Excel généré avec succès : {filename}")
print(f"   - {len(headers)} colonnes")
print(f"   - {row-1} lignes de données fictives")

